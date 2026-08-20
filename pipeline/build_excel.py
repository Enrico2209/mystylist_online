#!/usr/bin/env python3
"""
Esporta outfits_pool.jsonl in un workbook Excel navigabile.

Tre fogli:
  - "Outfit"    una riga per outfit (1653), con la foto di ogni capo, i titoli
                cliccabili verso la scheda prodotto e lo score
  - "Capi"      una riga per capo distinto (1801), con foto, brand, cluster
                stilistico e quante volte ricorre nel pool
  - "Riepilogo" statistiche calcolate con formule sul foglio Outfit

Uso:
    python3 build_excel.py
    python3 scripts/recalc.py outfits_pool.xlsx   # obbligatorio: le formule
                                                  # non hanno valori in cache
"""

import json
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image

from outfit_generation import classify_gender

from percorsi import DATI as BASE, CODICE, CATALOGO as _CATALOGO, IMMAGINI_SUPERATE, PROGETTI  # noqa: F401
ROOT = BASE / "nuvolari_full_organizzato"
JSONL = BASE / "outfits_pool.jsonl"
OUT_XLSX = BASE / "outfits_pool.xlsx"
TMP = BASE / ".xlsx_thumbs"

SLOTS = ["top", "bottom", "shoes", "outerwear", "accessory"]
SLOT_IT = {"top": "Top", "bottom": "Bottom", "shoes": "Scarpe",
           "outerwear": "Outerwear", "accessory": "Accessorio"}

THUMB_H = 60          # px: altezza miniatura
ROW_H = THUMB_H * 0.76  # punti: altezza riga che contiene la miniatura

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1C2620")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
CELL_FONT = Font(name=FONT, size=10)
LINK_FONT = Font(name=FONT, size=10, color="0563C1", underline="single")
NOTE_FONT = Font(name=FONT, size=9, italic=True, color="666F5F")


def make_thumb(rel_path: str, key: str) -> Path:
    """Miniatura JPEG (una per foto distinta, riusata in ogni cella che la mostra)."""
    out = TMP / f"{key}.jpg"
    if out.exists():
        return out
    with Image.open(ROOT / rel_path) as im:
        im = im.convert("RGB")
        w, h = im.size
        nh = THUMB_H
        nw = max(1, int(w * nh / h))
        im.resize((nw, nh), Image.LANCZOS).save(out, "JPEG", quality=72, optimize=True)
    return out


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"


def main():
    TMP.mkdir(exist_ok=True)
    rows = [json.loads(l) for l in open(JSONL, encoding="utf-8")]
    print(f"[*] {len(rows)} outfit da esportare")

    feats = pd.read_parquet(BASE / "features_clustered.parquet")
    feats = feats.set_index("relpath")

    wb = openpyxl.Workbook()

    # ---------------- foglio Outfit ----------------
    ws = wb.active
    ws.title = "Outfit"

    headers = (["N", "outfit_id", "Genere", "Score", "Ancora"]
               + [f"Foto {SLOT_IT[s]}" for s in SLOTS]
               + [SLOT_IT[s] for s in SLOTS]
               + ["N. capi"]
               + [f"id {SLOT_IT[s]}" for s in SLOTS])
    ws.append(headers)

    COL_FOTO0 = 6        # F
    COL_TITLE0 = 11      # K..O  (contigue: servono a COUNTIF nel foglio Capi)
    COL_NCAPI = 16       # P
    COL_PID0 = 17        # Q..U

    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=i - 1).font = CELL_FONT
        ws.cell(row=i, column=2, value=r["outfit_id"]).font = CELL_FONT
        ws.cell(row=i, column=3, value=r["gender"] or "n/d").font = CELL_FONT
        sc = ws.cell(row=i, column=4, value=r["outfit_score"])
        sc.font = CELL_FONT
        sc.number_format = "0.000"
        ws.cell(row=i, column=5, value=SLOT_IT.get(r["anchor_slot"], r["anchor_slot"])).font = CELL_FONT

        n_capi = 0
        for j, slot in enumerate(SLOTS):
            v = r["slots"].get(slot)
            if not v:
                continue
            n_capi += 1

            if v.get("display_image"):
                thumb = make_thumb(v["display_image"], v["product_id"])
                ws.add_image(XLImage(str(thumb)),
                             ws.cell(row=i, column=COL_FOTO0 + j).coordinate)

            t = ws.cell(row=i, column=COL_TITLE0 + j, value=v["title"])
            if v.get("url"):
                t.hyperlink = v["url"]
                t.font = LINK_FONT
            else:
                t.font = CELL_FONT
            t.alignment = Alignment(vertical="center", wrap_text=False)

            ws.cell(row=i, column=COL_PID0 + j, value=v["product_id"]).font = CELL_FONT

        ws.cell(row=i, column=COL_NCAPI, value=n_capi).font = CELL_FONT
        ws.row_dimensions[i].height = ROW_H

        if (i - 1) % 250 == 0:
            print(f"    outfit {i - 1}/{len(rows)}")

    style_header(ws, len(headers))
    ws.auto_filter.ref = f"A1:P{len(rows) + 1}"
    for c, w in [(1, 6), (2, 14), (3, 9), (4, 8), (5, 11)]:
        ws.column_dimensions[get_column_letter(c)].width = w
    for j in range(len(SLOTS)):
        ws.column_dimensions[get_column_letter(COL_FOTO0 + j)].width = 7.5
        ws.column_dimensions[get_column_letter(COL_TITLE0 + j)].width = 34
        # gli id prodotto servono alle formule del foglio Capi, non alla lettura
        ws.column_dimensions[get_column_letter(COL_PID0 + j)].hidden = True
    ws.column_dimensions[get_column_letter(COL_NCAPI)].width = 8

    last = len(rows) + 1
    pid_range = f"Outfit!${get_column_letter(COL_PID0)}$2:${get_column_letter(COL_PID0 + 4)}${last}"

    # ---------------- foglio Capi ----------------
    ws2 = wb.create_sheet("Capi")
    ws2.append(["product_id", "Foto", "Capo", "Slot", "Brand", "Genere",
                "Cluster stile", "Formalità", "N. outfit", "N. foto", "Percorso"])

    capi = {}
    for r in rows:
        for slot, v in r["slots"].items():
            if v and v["relpath"] not in capi:
                capi[v["relpath"]] = (slot, v)

    for i, (relpath, (slot, v)) in enumerate(sorted(capi.items()), start=2):
        ws2.cell(row=i, column=1, value=v["product_id"]).font = CELL_FONT
        if v.get("display_image"):
            ws2.add_image(XLImage(str(make_thumb(v["display_image"], v["product_id"]))),
                          ws2.cell(row=i, column=2).coordinate)
        t = ws2.cell(row=i, column=3, value=v["title"])
        if v.get("url"):
            t.hyperlink = v["url"]
            t.font = LINK_FONT
        else:
            t.font = CELL_FONT
        ws2.cell(row=i, column=4, value=SLOT_IT.get(slot, slot)).font = CELL_FONT

        row_f = feats.loc[relpath] if relpath in feats.index else None
        ws2.cell(row=i, column=5, value=(row_f["brand_slug"] if row_f is not None else None)).font = CELL_FONT
        ws2.cell(row=i, column=6, value=(classify_gender(v["title"], relpath) or "n/d")).font = CELL_FONT
        cl = ws2.cell(row=i, column=7,
                      value=(int(row_f["style_cluster"]) if row_f is not None else None))
        cl.font = CELL_FONT
        fm = ws2.cell(row=i, column=8,
                      value=(float(row_f["formality_norm"]) if row_f is not None else None))
        fm.font = CELL_FONT
        fm.number_format = "0.00"

        # conteggio ricavato dal foglio Outfit, non precalcolato in Python:
        # resta corretto anche se si filtrano o modificano le righe a mano
        cnt = ws2.cell(row=i, column=9, value=f'=COUNTIF({pid_range},A{i})')
        cnt.font = CELL_FONT
        ws2.cell(row=i, column=10, value=len(v.get("all_images") or [])).font = CELL_FONT
        ws2.cell(row=i, column=11, value=relpath).font = CELL_FONT
        ws2.row_dimensions[i].height = ROW_H

    style_header(ws2, 11)
    ws2.auto_filter.ref = f"A1:K{len(capi) + 1}"
    for c, w in [(1, 14), (2, 7.5), (3, 44), (4, 12), (5, 18), (6, 9),
                 (7, 12), (8, 10), (9, 10), (10, 9), (11, 52)]:
        ws2.column_dimensions[get_column_letter(c)].width = w

    # ---------------- foglio Riepilogo ----------------
    ws3 = wb.create_sheet("Riepilogo", 0)
    ws3.append(["Metrica", "Valore", "Note"])
    metriche = [
        ("Outfit totali", f"=COUNTA(Outfit!$B$2:$B${last})", "Combinazioni uniche nel pool"),
        ("Capi distinti usati", f"=COUNTA(Capi!$A$2:$A${len(capi) + 1})", "Prodotti che compaiono in almeno un outfit"),
        ("Score minimo", f"=MIN(Outfit!$D$2:$D${last})", "Soglia impostata alla generazione: 0,60"),
        ("Score massimo", f"=MAX(Outfit!$D$2:$D${last})", ""),
        ("Score medio", f"=AVERAGE(Outfit!$D$2:$D${last})", ""),
        ("Outfit uomo", f'=COUNTIF(Outfit!$C$2:$C${last},"uomo")', ""),
        ("Outfit donna", f'=COUNTIF(Outfit!$C$2:$C${last},"donna")', ""),
        ("Outfit genere n/d", f'=COUNTIF(Outfit!$C$2:$C${last},"n/d")', "Genere non deducibile da titolo o categoria"),
        ("Ancora: Top", f'=COUNTIF(Outfit!$E$2:$E${last},"Top")', ""),
        ("Ancora: Bottom", f'=COUNTIF(Outfit!$E$2:$E${last},"Bottom")', ""),
        ("Ancora: Scarpe", f'=COUNTIF(Outfit!$E$2:$E${last},"Scarpe")', ""),
        ("Outfit con outerwear", f'=COUNTA(Outfit!$N$2:$N${last})', "Slot opzionale, aggiunto solo se sopra soglia"),
        ("Outfit con accessorio", f'=COUNTA(Outfit!$O$2:$O${last})', "Slot opzionale, aggiunto solo se sopra soglia"),
    ]
    for k, (nome, formula, nota) in enumerate(metriche, start=2):
        ws3.cell(row=k, column=1, value=nome).font = CELL_FONT
        c = ws3.cell(row=k, column=2, value=formula)
        c.font = CELL_FONT
        c.number_format = "0.000" if "Score" in nome else "0"
        ws3.cell(row=k, column=3, value=nota).font = NOTE_FONT

    r0 = len(metriche) + 3
    ws3.cell(row=r0, column=1, value="Come leggere lo Score").font = Font(name=FONT, bold=True, size=10)
    for off, txt in enumerate([
        "È il MINIMO dei punteggi di compatibilità fra tutte le coppie di capi dell'outfit, non la media:",
        "un solo abbinamento debole abbassa l'intero outfit invece di essere compensato dagli altri.",
        "Ogni punteggio di coppia combina armonia di colore (spazio Lab) e affinità di stile (coseno fra i vettori).",
        "Generato da run_outfit_pipeline() con soglia minima 0,60 — vedi outfit_generation.py.",
    ]):
        ws3.cell(row=r0 + 1 + off, column=1, value=txt).font = NOTE_FONT

    style_header(ws3, 3)
    ws3.freeze_panes = "A2"
    for c, w in [(1, 26), (2, 14), (3, 62)]:
        ws3.column_dimensions[get_column_letter(c)].width = w

    wb.save(OUT_XLSX)
    print(f"[OK] scritto {OUT_XLSX}")


if __name__ == "__main__":
    main()
